"""组织管理相关API"""

import base64
import csv
import io
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, Header, HTTPException, Response, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_super_admin, get_current_user_management_user
from app.model import Department, DepartmentMember, Project, ProjectDepartment, User
from app.rbac import (
    can_access_user_management,
    get_primary_platform_role,
    is_ordinary_admin,
    is_super_admin,
)
from app.schema import (
    DepartmentCreate, DepartmentUpdate, DepartmentResponse,
    DepartmentMemberCreate, DepartmentMemberUpdate, DepartmentMemberResponse,
    DepartmentMemberImportCommitResponse, DepartmentMemberImportNormalizedRow,
    DepartmentMemberImportPreviewResponse, DepartmentMemberImportRequest,
    DepartmentMemberImportRowResult,
    ProjectCreate, ProjectUpdate, ProjectResponse, ProjectDetailResponse,
    Message, UserPermissionInfo,
    UserDepartmentProjectListResponse, UserDepartmentProjectResponse
)
from app.service.project import get_project_service, ProjectServiceError

router = APIRouter(prefix="/org", tags=["organization"])

logger = logging.getLogger(__name__)

DEPARTMENT_MEMBER_IMPORT_REQUIRED_HEADERS = {"username"}
DEPARTMENT_MEMBER_IMPORT_ALLOWED_HEADERS = {"username", "role"}
DEPARTMENT_MEMBER_IMPORT_ALLOWED_ROLES = {"leader", "vice_leader", "member"}
DEPARTMENT_MEMBER_IMPORT_ALLOWED_MODES = {"skip_existing", "update_role"}
DEPARTMENT_MEMBER_IMPORT_TEMPLATE_FILENAME = "secflow-department-member-import-template.xlsx"


def log_access_denied(user: User, resource_type: str, resource_id: int, reason: str):
    """记录访问被拒绝的日志"""
    logger.warning(
        f"访问被拒绝 - 用户: {user.username}(ID:{user.id}), "
        f"资源类型: {resource_type}, 资源ID: {resource_id}, "
        f"原因: {reason}"
    )


def is_admin(user: User) -> bool:
    """检查用户是否为管理员"""
    return is_super_admin(user)


def get_user_department_ids(db: Session, user_id: int) -> List[int]:
    """获取用户所属的所有部门ID"""
    memberships = db.query(DepartmentMember.department_id).filter(
        DepartmentMember.user_id == user_id
    ).all()
    return [m.department_id for m in memberships]


def get_all_descendant_ids(db: Session, department_ids: List[int]) -> List[int]:
    """获取所有下级部门ID（递归）"""
    all_descendants = set(department_ids)
    
    def find_children(parent_ids: List[int]):
        children = db.query(Department.id).filter(
            Department.parent_id.in_(parent_ids)
        ).all()
        child_ids = [c.id for c in children]
        if child_ids:
            new_ids = [cid for cid in child_ids if cid not in all_descendants]
            if new_ids:
                all_descendants.update(new_ids)
                find_children(new_ids)
    
    find_children(department_ids)
    return list(all_descendants)


def get_accessible_department_ids(db: Session, user: User) -> List[int]:
    """获取用户可访问的所有部门ID。"""
    if is_super_admin(user):
        return None

    if not is_ordinary_admin(user):
        return []

    user_dept_ids = get_user_department_ids(db, user.id)
    if not user_dept_ids:
        return []

    return get_all_descendant_ids(db, user_dept_ids)


def get_manageable_department_ids(db: Session, user: User) -> List[int]:
    """获取用户可管理的部门ID。普通管理员仅能管理所属部门及下级部门。"""
    if is_super_admin(user):
        return None

    if not is_ordinary_admin(user):
        return []

    user_dept_ids = get_user_department_ids(db, user.id)
    if not user_dept_ids:
        return []

    return get_all_descendant_ids(db, user_dept_ids)


def get_department_structure_manageable_ids(db: Session, user: User) -> List[int]:
    """仅超级管理员可以管理部门结构。"""
    if is_super_admin(user):
        return None
    return []


def can_move_member_between_departments(
    db: Session,
    current_user: User,
    target_member: DepartmentMember,
    new_department_id: int,
) -> tuple[bool, str]:
    """普通管理员仅能移动其部门树内的成员。"""
    if is_super_admin(current_user):
        return True, ""

    manageable_ids = get_manageable_department_ids(db, current_user) or []
    if target_member.department_id not in manageable_ids:
        return False, "无权调整该用户的所属部门"

    if new_department_id not in manageable_ids:
        return False, "目标部门不在可管理范围内"

    return True, ""


def ensure_department_member_operation_scope(
    db: Session,
    current_user: User,
    member: DepartmentMember
) -> DepartmentMember:
    """确保当前用户可以操作目标成员。"""
    if is_super_admin(current_user):
        return member

    manageable_ids = get_manageable_department_ids(db, current_user) or []
    if member.department_id not in manageable_ids:
        log_access_denied(current_user, "部门成员", member.department_id, "目标成员不在可管理范围内")
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="目标成员不在可管理范围内"
        )
    return member


def ensure_department_member_management_scope(
    db: Session,
    current_user: User,
    department_id: int
) -> Department:
    """确保当前用户可以管理目标部门成员。"""
    department = db.query(Department).filter(Department.id == department_id).first()
    if not department:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="部门不存在"
        )

    if is_super_admin(current_user):
        return department

    manageable_ids = get_manageable_department_ids(db, current_user) or []
    if department_id not in manageable_ids:
        log_access_denied(current_user, "部门成员", department_id, "目标部门不在可管理范围内")
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="目标部门不在可管理范围内"
        )
    return department


def get_import_allowed_roles(current_user: User) -> set[str]:
    if is_super_admin(current_user):
        return set(DEPARTMENT_MEMBER_IMPORT_ALLOWED_ROLES)
    return {"member"}


def normalize_import_csv(content: str) -> str:
    return (content or "").replace("\r\n", "\n").replace("\r", "\n").lstrip("\ufeff")


def _validate_import_headers(headers: List[str], file_label: str) -> None:
    if not headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{file_label} 缺少表头"
        )

    missing_headers = DEPARTMENT_MEMBER_IMPORT_REQUIRED_HEADERS.difference(headers)
    if missing_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{file_label} 缺少必填列: {', '.join(sorted(missing_headers))}"
        )

    unsupported_headers = [header for header in headers if header not in DEPARTMENT_MEMBER_IMPORT_ALLOWED_HEADERS]
    if unsupported_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{file_label} 存在不支持的列: {', '.join(unsupported_headers)}"
        )


def _load_department_member_import_rows_from_csv(csv_content: str) -> List[Tuple[int, Dict[str, str]]]:
    normalized = normalize_import_csv(csv_content)
    if not normalized.strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV 内容不能为空"
        )

    reader = csv.DictReader(io.StringIO(normalized))
    headers = [header.strip() for header in (reader.fieldnames or []) if header and header.strip()]
    _validate_import_headers(headers, "CSV")

    rows: List[Tuple[int, Dict[str, str]]] = []
    for row_no, row in enumerate(reader, start=2):
        normalized_row = {
            key.strip(): (value.strip() if isinstance(value, str) else "")
            for key, value in (row or {}).items()
            if key is not None and key.strip()
        }
        if not any(normalized_row.values()):
            continue
        rows.append((row_no, normalized_row))

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV 不包含有效数据行"
        )
    return rows


def _load_department_member_import_rows_from_excel(file_bytes: bytes) -> List[Tuple[int, Dict[str, str]]]:
    try:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Excel 文件解析失败: {exc}"
        ) from exc

    sheet = workbook.active
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [str(cell).strip() for cell in (header_cells or []) if str(cell or "").strip()]
    _validate_import_headers(headers, "Excel")

    rows: List[Tuple[int, Dict[str, str]]] = []
    for row_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        normalized_row: Dict[str, str] = {}
        has_value = False
        for index, header in enumerate(headers):
            cell_value = values[index] if values and index < len(values) else ""
            value = str(cell_value).strip() if cell_value is not None else ""
            normalized_row[header] = value
            if value:
                has_value = True
        if not has_value:
            continue
        rows.append((row_no, normalized_row))

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel 不包含有效数据行"
        )
    return rows


def load_department_member_import_rows(request: DepartmentMemberImportRequest) -> List[Tuple[int, Dict[str, str]]]:
    if request.file_content_base64:
        if not request.filename:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel 导入缺少文件名"
            )
        try:
            file_bytes = base64.b64decode(request.file_content_base64)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"导入文件内容解码失败: {exc}"
            ) from exc

        suffix = Path(request.filename).suffix.lower()
        if suffix == ".csv":
            try:
                csv_content = file_bytes.decode("utf-8-sig")
            except UnicodeDecodeError:
                csv_content = file_bytes.decode("gbk")
            return _load_department_member_import_rows_from_csv(csv_content)
        if suffix == ".xlsx":
            return _load_department_member_import_rows_from_excel(file_bytes)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="当前仅支持导入 .xlsx 或 .csv 文件"
        )

    if request.csv_content is not None:
        return _load_department_member_import_rows_from_csv(request.csv_content)

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="请上传 Excel/CSV 文件，或提供 CSV 内容"
    )


def build_department_member_import_template_workbook() -> bytes:
    workbook = Workbook()
    template_sheet = workbook.active
    template_sheet.title = "成员导入模板"
    template_sheet.append(["username", "role"])
    template_sheet.append(["", ""])
    template_sheet.append(["", ""])
    template_sheet.freeze_panes = "A2"
    template_sheet.column_dimensions["A"].width = 28
    template_sheet.column_dimensions["B"].width = 18
    for cell in template_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="2563EB")

    guide_sheet = workbook.create_sheet("填写说明")
    guide_sheet.column_dimensions["A"].width = 22
    guide_sheet.column_dimensions["B"].width = 90
    guide_sheet.append(["项目", "说明"])
    guide_sheet.append(["username", "必填，填写系统里已经存在的用户名，例如 zhangsan"])
    guide_sheet.append(["role", "可选，可填写 member / vice_leader / leader；普通管理员只允许 member"])
    guide_sheet.append(["目标部门", "不用填在文件里，系统会自动导入到你当前选中的部门"])
    guide_sheet.append(["导入模式", "已存在则跳过；超级管理员还可以选择已存在则更新角色"])
    guide_sheet.append(["注意事项", "每个部门同一时间只能有一个 leader；文件第一行必须是表头"])
    for cell in guide_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="0F766E")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def preview_department_member_import(
    db: Session,
    current_user: User,
    request: DepartmentMemberImportRequest
) -> DepartmentMemberImportPreviewResponse:
    if request.mode not in DEPARTMENT_MEMBER_IMPORT_ALLOWED_MODES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="mode 仅支持 skip_existing 或 update_role"
        )

    target_department = ensure_department_member_management_scope(db, current_user, request.department_id)
    if not is_super_admin(current_user) and request.mode == "update_role":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="普通管理员不能通过导入批量修改角色"
        )

    rows = load_department_member_import_rows(request)
    allowed_roles = get_import_allowed_roles(current_user)
    username_counts: Dict[str, int] = {}
    leader_rows = 0

    for _row_no, raw_row in rows:
        username = (raw_row.get("username") or "").strip()
        if username:
            username_counts[username] = username_counts.get(username, 0) + 1
        role = (raw_row.get("role") or "").strip().lower() or "member"
        if role == "leader":
            leader_rows += 1

    current_leader = db.query(DepartmentMember).filter(
        DepartmentMember.department_id == target_department.id,
        DepartmentMember.role == "leader"
    ).first()

    results: List[DepartmentMemberImportRowResult] = []
    for row_no, raw_row in rows:
        messages: List[str] = []
        errors: List[str] = []
        username = (raw_row.get("username") or "").strip()
        role = (raw_row.get("role") or "").strip().lower() or "member"

        if not username:
            errors.append("用户名不能为空")
        elif username_counts.get(username, 0) > 1:
            errors.append("CSV 文件内用户名重复")

        if role not in DEPARTMENT_MEMBER_IMPORT_ALLOWED_ROLES:
            errors.append("role 仅支持 leader、vice_leader、member")
        elif role not in allowed_roles:
            errors.append("当前角色权限仅允许导入普通成员")

        user = db.query(User).filter(User.username == username).first() if username else None
        if username and not user:
            errors.append("用户不存在，请先创建账号")

        existing_member = None
        first_membership = None
        if user:
            existing_member = db.query(DepartmentMember).filter(
                DepartmentMember.user_id == user.id,
                DepartmentMember.department_id == target_department.id
            ).first()
            first_membership = db.query(DepartmentMember).filter(
                DepartmentMember.user_id == user.id
            ).order_by(DepartmentMember.id.asc()).first()

        action = "create"
        if existing_member:
            if request.mode == "skip_existing":
                action = "skip_existing"
                messages.append("用户已在该部门，执行时将跳过")
            elif request.mode == "update_role":
                action = "update_role"
                messages.append("用户已在该部门，执行时将更新角色")

        if role == "leader":
            if leader_rows > 1:
                errors.append("同一批导入中只能有一个组长")
            if current_leader:
                if not existing_member or current_leader.id != existing_member.id:
                    errors.append("该部门已经有组长")

        normalized = DepartmentMemberImportNormalizedRow(
            username=username,
            department_id=target_department.id,
            department_name=target_department.name,
            role=role,
            action=action,
            existing_member_id=existing_member.id if existing_member else None,
            existing_department_id=first_membership.department_id if first_membership else None,
            existing_department_name=first_membership.department.name if first_membership and first_membership.department else None,
        )
        status_value = "valid" if not errors else "error"

        results.append(DepartmentMemberImportRowResult(
            row_no=row_no,
            username=username,
            status=status_value,
            messages=errors + messages,
            normalized=normalized,
            member_id=existing_member.id if existing_member else None,
        ))

    valid_rows = sum(1 for item in results if item.status == "valid")
    return DepartmentMemberImportPreviewResponse(
        total_rows=len(results),
        valid_rows=valid_rows,
        error_rows=len(results) - valid_rows,
        rows=results,
    )


def get_project_department_ids(db: Session, project_id: int) -> List[int]:
    """获取项目绑定的部门ID列表。"""
    rows = db.query(ProjectDepartment.department_id).filter(
        ProjectDepartment.project_id == project_id
    ).all()
    return [row.department_id for row in rows]


def get_project_departments(db: Session, project_id: int) -> List[Department]:
    """获取项目绑定的部门列表。"""
    return db.query(Department).join(ProjectDepartment).filter(
        ProjectDepartment.project_id == project_id
    ).all()


def is_department_scope_allowed(allowed_department_ids: Optional[List[int]], target_department_ids: List[int]) -> bool:
    """检查目标部门是否全部在允许范围内。"""
    if allowed_department_ids is None:
        return True

    if not target_department_ids:
        return False

    allowed_id_set = set(allowed_department_ids)
    return all(department_id in allowed_id_set for department_id in target_department_ids)


def can_manage_org_project(
    db: Session,
    current_user: User,
    project: Project,
) -> tuple[bool, str]:
    """普通管理员仅能管理绑定在自己部门树内的项目。"""
    if is_super_admin(current_user):
        return True, ""

    if not is_ordinary_admin(current_user):
        return False, "您没有权限管理项目"

    manageable_ids = get_manageable_department_ids(db, current_user) or []
    project_department_ids = get_project_department_ids(db, project.id)

    if not project_department_ids:
        return False, "普通管理员只能管理已绑定所属部门或下级部门的项目"

    if not is_department_scope_allowed(manageable_ids, project_department_ids):
        return False, "无权管理上级部门或其他部门的项目"

    return True, ""


def validate_project_assignment_scope(
    db: Session,
    current_user: User,
    department_ids: List[int],
    is_public: bool,
) -> tuple[bool, str]:
    """校验项目绑定部门是否在当前用户可管理范围内。"""
    if is_super_admin(current_user):
        return True, ""

    if not is_ordinary_admin(current_user):
        return False, "您没有权限管理项目"

    if is_public:
        return False, "普通管理员只能创建或维护绑定所属部门及下级部门的私有项目"

    if not department_ids:
        return False, "普通管理员必须为项目绑定所属部门或下级部门"

    manageable_ids = get_manageable_department_ids(db, current_user) or []
    if not is_department_scope_allowed(manageable_ids, department_ids):
        return False, "只能选择所属部门及下级部门"

    return True, ""


def build_project_detail_response(db: Session, project: Project) -> ProjectDetailResponse:
    """构建包含部门信息的项目响应。"""
    from app.schema import DepartmentResponse

    departments = get_project_departments(db, project.id)
    return ProjectDetailResponse(
        id=project.id,
        name=project.name,
        description=project.description,
        is_public=project.is_public,
        created_at=project.created_at,
        updated_at=project.updated_at,
        departments=[DepartmentResponse.from_orm(dept) for dept in departments]
    )


def check_circular_reference(db: Session, department_id: int, new_parent_id: int) -> bool:
    """检查是否存在循环引用"""
    visited = set()
    current_id = new_parent_id
    
    while current_id is not None:
        if current_id == department_id:
            return True
        if current_id in visited:
            return True
        visited.add(current_id)
        
        parent_dept = db.query(Department).filter(Department.id == current_id).first()
        if not parent_dept:
            break
        current_id = parent_dept.parent_id
    
    return False


# ============ 部门管理 ============

@router.get("/user-permissions", response_model=UserPermissionInfo)
def get_user_permissions(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """获取当前用户的权限信息"""
    is_user_admin = is_admin(current_user)
    user_dept_ids = get_user_department_ids(db, current_user.id)
    manageable_dept_ids = get_manageable_department_ids(db, current_user)
    dept_structure_ids = get_department_structure_manageable_ids(db, current_user)
    role_names = current_user.get_all_role_names()

    return UserPermissionInfo(
        user_id=current_user.id,
        is_admin=is_user_admin,
        platform_role=get_primary_platform_role(current_user),
        department_ids=user_dept_ids,
        manageable_department_ids=manageable_dept_ids if manageable_dept_ids is not None else [],
        department_structure_manageable_ids=dept_structure_ids if dept_structure_ids is not None else [],
        role_names=role_names,
        can_access_user_management=can_access_user_management(current_user),
        can_manage_users=is_super_admin(current_user),
        can_manage_roles=is_super_admin(current_user),
        can_manage_departments=is_super_admin(current_user),
        can_manage_department_members=is_super_admin(current_user) or is_ordinary_admin(current_user),
        can_manage_org_projects=is_super_admin(current_user) or is_ordinary_admin(current_user),
    )


def get_user_department_info(db: Session, user_id: int) -> tuple[Optional[int], Optional[str]]:
    """
    获取用户的主要部门信息

    Returns:
        (部门ID, 部门名称)
    """
    member = db.query(DepartmentMember).filter(
        DepartmentMember.user_id == user_id
    ).first()

    if not member:
        return None, None

    department = db.query(Department).filter(Department.id == member.department_id).first()
    if not department:
        return None, None

    return department.id, department.name


@router.get("/user-department-projects", response_model=UserDepartmentProjectListResponse)
def get_user_department_projects(
    authorization: str = Header(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """
    获取用户可见的项目列表

    - 公开项目：所有人可见
    - 非公开项目：仅项目归属部门在当前用户可访问部门范围内时可见
    - 返回这些项目的信息，包括归属部门信息
    """
    if authorization is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="缺少认证信息"
        )

    accessible_dept_ids = get_accessible_department_ids(db, current_user)

    try:
        # 调用project服务获取所有项目
        project_service = get_project_service()
        projects_data = project_service.get_all_projects(authorization.replace("Bearer ", ""))
    except ProjectServiceError as e:
        logger.error(f"获取项目列表失败: {e}")
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=f"无法获取项目列表: {str(e)}"
        )

    # 筛选项目
    result_projects = []
    for proj in projects_data:
        is_public = proj.get("is_public", False)
        owner_id_str = proj.get("owner_id")
        project_department_id = proj.get("department_id")
        project_department_name = proj.get("department_name")
        can_manage = bool(proj.get("can_manage", False))

        owner_dept_id = None
        owner_dept_name = None
        if owner_id_str:
            try:
                owner_id = int(owner_id_str)
                owner_dept_id, owner_dept_name = get_user_department_info(db, owner_id)
            except (ValueError, TypeError):
                pass

        # 公开项目：所有人可见
        if is_public:
            roles = proj.get("roles", [])
            result_projects.append(UserDepartmentProjectResponse(
                id=proj.get("id"),
                name=proj.get("name"),
                description=proj.get("description"),
                owner_id=owner_id_str or "",
                owner_name=proj.get("owner_name"),
                k8s_namespace=proj.get("k8s_namespace"),
                status=proj.get("status"),
                is_public=is_public,
                department_id=project_department_id,
                department_name=project_department_name,
                can_manage=can_manage,
                created_at=proj.get("created_at"),
                updated_at=proj.get("updated_at"),
                roles=roles,
                owner_department_id=owner_dept_id,
                owner_department_name=owner_dept_name
            ))
        else:
            # 非公开项目：仅项目归属部门在当前用户可访问部门范围内时可见
            effective_department_id = project_department_id or owner_dept_id
            effective_department_name = project_department_name or owner_dept_name

            if effective_department_id is None:
                continue
            if accessible_dept_ids is None or effective_department_id in accessible_dept_ids:
                roles = proj.get("roles", [])
                result_projects.append(UserDepartmentProjectResponse(
                    id=proj.get("id"),
                    name=proj.get("name"),
                    description=proj.get("description"),
                    owner_id=owner_id_str or "",
                    owner_name=proj.get("owner_name"),
                    k8s_namespace=proj.get("k8s_namespace"),
                    status=proj.get("status"),
                    is_public=is_public,
                    department_id=effective_department_id,
                    department_name=effective_department_name,
                    can_manage=can_manage,
                    created_at=proj.get("created_at"),
                    updated_at=proj.get("updated_at"),
                    roles=roles,
                    owner_department_id=owner_dept_id,
                    owner_department_name=owner_dept_name
                ))

    return UserDepartmentProjectListResponse(
        total=len(result_projects),
        projects=result_projects
    )


@router.post("/departments", response_model=DepartmentResponse, status_code=status.HTTP_201_CREATED)
def create_department(
    department: DepartmentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """创建部门。仅超级管理员可用。"""
    # 验证父部门是否存在
    if department.parent_id is not None and department.parent_id != 0:
        parent_dept = db.query(Department).filter(Department.id == department.parent_id).first()
        if not parent_dept:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="父部门不存在"
            )
        
    else:
        department.parent_id = None
    
    # 创建部门
    db_department = Department(
        name=department.name,
        description=department.description,
        parent_id=department.parent_id if department.parent_id != 0 else None
    )
    db.add(db_department)
    db.commit()
    db.refresh(db_department)
    
    logger.info(f"创建部门: 用户{current_user.username}创建了部门{db_department.name}(ID: {db_department.id})")
    
    return db_department


@router.get("/departments", response_model=List[DepartmentResponse])
def get_departments(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """获取部门列表（仅返回用户可访问的部门）"""
    accessible_ids = get_accessible_department_ids(db, current_user)
    
    if accessible_ids is None:
        departments = db.query(Department).all()
    else:
        departments = db.query(Department).filter(Department.id.in_(accessible_ids)).all()
    
    return departments


@router.get("/departments/{department_id}", response_model=DepartmentResponse)
def get_department(
    department_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """获取部门详情"""
    accessible_ids = get_accessible_department_ids(db, current_user)
    
    if accessible_ids is not None and department_id not in accessible_ids:
        log_access_denied(current_user, "部门", department_id, "用户无权访问该部门（非所属部门或下级部门）")
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="您没有权限访问该部门"
        )
    
    department = db.query(Department).filter(Department.id == department_id).first()
    if not department:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="部门不存在"
        )
    return department


@router.put("/departments/{department_id}", response_model=DepartmentResponse)
def update_department(
    department_id: int,
    department: DepartmentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """更新部门信息。仅超级管理员可用。"""
    # 验证部门是否存在
    db_department = db.query(Department).filter(Department.id == department_id).first()
    if not db_department:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="部门不存在"
        )
    
    # 检查用户是否有权限管理该部门
    # 验证父部门是否存在
    if department.parent_id is not None:
        # 不能将自己设置为父部门
        if department.parent_id == department_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="不能将自己设置为父部门"
            )
        
        if department.parent_id != 0:
            parent_dept = db.query(Department).filter(Department.id == department.parent_id).first()
            if not parent_dept:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="父部门不存在"
                )
            
            # 检查循环引用
            if check_circular_reference(db, department_id, department.parent_id):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="不能设置循环引用的父部门，这会形成闭环关系"
                )
    
    # 更新部门信息
    if department.name is not None:
        db_department.name = department.name
    if department.description is not None:
        db_department.description = department.description
    if department.parent_id is not None:
        db_department.parent_id = department.parent_id if department.parent_id != 0 else None
    
    db.commit()
    db.refresh(db_department)
    return db_department


@router.delete("/departments/{department_id}", response_model=Message)
def delete_department(
    department_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """删除部门。仅超级管理员可用。"""
    # 验证部门是否存在
    db_department = db.query(Department).filter(Department.id == department_id).first()
    if not db_department:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="部门不存在"
        )
    
    # 检查是否有子部门
    if db_department.children:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="该部门下有子部门，无法删除"
        )
    
    # 删除部门成员
    db.query(DepartmentMember).filter(DepartmentMember.department_id == department_id).delete()
    
    # 删除部门与项目的关联
    db.query(ProjectDepartment).filter(ProjectDepartment.department_id == department_id).delete()
    
    # 删除部门
    db.delete(db_department)
    db.commit()
    
    return Message(message="部门删除成功")


# ============ 部门成员管理 ============

@router.get("/department-members/import/template")
def download_department_member_import_template(
    current_user: User = Depends(get_current_user_management_user)
):
    """下载部门成员导入模板。"""
    if not (is_super_admin(current_user) or is_ordinary_admin(current_user)):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="当前用户无权导入部门成员"
        )
    return Response(
        content=build_department_member_import_template_workbook(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{DEPARTMENT_MEMBER_IMPORT_TEMPLATE_FILENAME}"'
        }
    )


@router.post("/department-members/import/preview", response_model=DepartmentMemberImportPreviewResponse)
def preview_department_members_import(
    request: DepartmentMemberImportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """预校验部门成员导入。"""
    return preview_department_member_import(db, current_user, request)


@router.post("/department-members/import/commit", response_model=DepartmentMemberImportCommitResponse)
def commit_department_members_import(
    request: DepartmentMemberImportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """执行部门成员导入。按行提交，单行失败不影响整批。"""
    preview = preview_department_member_import(db, current_user, request)
    success_rows = 0
    skipped_rows = 0
    results: List[DepartmentMemberImportRowResult] = []

    for row in preview.rows:
        if row.status != "valid" or row.normalized is None:
            results.append(row)
            continue

        normalized = row.normalized
        try:
            if normalized.action == "skip_existing":
                skipped_rows += 1
                results.append(DepartmentMemberImportRowResult(
                    row_no=row.row_no,
                    username=normalized.username,
                    status="skipped",
                    messages=["成员已存在，已跳过"],
                    normalized=normalized,
                    member_id=normalized.existing_member_id,
                ))
                continue

            user = db.query(User).filter(User.username == normalized.username).first()
            department = ensure_department_member_management_scope(db, current_user, normalized.department_id)
            if not user:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="用户不存在，请先创建账号"
                )

            if normalized.role == "leader":
                existing_leader = db.query(DepartmentMember).filter(
                    DepartmentMember.department_id == department.id,
                    DepartmentMember.role == "leader"
                ).first()
                if existing_leader and existing_leader.id != normalized.existing_member_id:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="该部门已经有组长"
                    )

            if normalized.action == "update_role":
                if not is_super_admin(current_user):
                    raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail="普通管理员不能通过导入批量修改角色"
                    )
                db_member = db.query(DepartmentMember).filter(
                    DepartmentMember.id == normalized.existing_member_id
                ).first()
                if not db_member:
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND,
                        detail="部门成员不存在"
                    )
                db_member.role = normalized.role
                db.commit()
                db.refresh(db_member)
                success_rows += 1
                results.append(DepartmentMemberImportRowResult(
                    row_no=row.row_no,
                    username=normalized.username,
                    status="success",
                    messages=["成员角色更新成功"],
                    normalized=normalized,
                    member_id=db_member.id,
                ))
                continue

            existing_member = db.query(DepartmentMember).filter(
                DepartmentMember.user_id == user.id,
                DepartmentMember.department_id == department.id
            ).first()
            if existing_member:
                skipped_rows += 1
                results.append(DepartmentMemberImportRowResult(
                    row_no=row.row_no,
                    username=normalized.username,
                    status="skipped",
                    messages=["成员已存在，已跳过"],
                    normalized=normalized,
                    member_id=existing_member.id,
                ))
                continue

            db_member = DepartmentMember(
                user_id=user.id,
                department_id=department.id,
                role=normalized.role
            )
            db.add(db_member)
            db.commit()
            db.refresh(db_member)
            success_rows += 1
            results.append(DepartmentMemberImportRowResult(
                row_no=row.row_no,
                username=normalized.username,
                status="success",
                messages=["成员导入成功"],
                normalized=normalized,
                member_id=db_member.id,
            ))
        except HTTPException as exc:
            db.rollback()
            results.append(DepartmentMemberImportRowResult(
                row_no=row.row_no,
                username=normalized.username,
                status="error",
                messages=[str(exc.detail)],
                normalized=normalized,
            ))
        except Exception as exc:
            db.rollback()
            results.append(DepartmentMemberImportRowResult(
                row_no=row.row_no,
                username=normalized.username,
                status="error",
                messages=[f"导入失败: {str(exc)}"],
                normalized=normalized,
            ))

    return DepartmentMemberImportCommitResponse(
        total_rows=len(results),
        success_rows=success_rows,
        skipped_rows=skipped_rows,
        failed_rows=len([item for item in results if item.status == "error"]),
        rows=results,
    )

@router.post("/department-members", response_model=DepartmentMemberResponse, status_code=status.HTTP_201_CREATED)
def add_department_member(
    member: DepartmentMemberCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """添加部门成员。普通管理员可管理所属部门及下级部门的成员。"""
    department = ensure_department_member_management_scope(db, current_user, member.department_id)
    
    # 验证用户是否存在
    user = db.query(User).filter(User.id == member.user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )
    
    # 验证用户是否已经是部门成员
    existing_member = db.query(DepartmentMember).filter(
        DepartmentMember.user_id == member.user_id,
        DepartmentMember.department_id == member.department_id
    ).first()
    if existing_member:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="用户已经是该部门的成员"
        )
    
    # 验证是否已经有组长
    if member.role == "leader":
        existing_leader = db.query(DepartmentMember).filter(
            DepartmentMember.department_id == member.department_id,
            DepartmentMember.role == "leader"
        ).first()
        if existing_leader:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="该部门已经有组长"
            )
    
    # 添加成员
    db_member = DepartmentMember(
        user_id=member.user_id,
        department_id=member.department_id,
        role=member.role
    )
    db.add(db_member)
    db.commit()
    db.refresh(db_member)
    
    logger.info(f"添加部门成员: 用户{current_user.username}将用户{user.username}添加到部门{department.name}，角色: {member.role}")
    
    # 构建响应
    response = DepartmentMemberResponse(
        id=db_member.id,
        user_id=db_member.user_id,
        username=user.username,
        department_id=db_member.department_id,
        department_name=department.name,
        role=db_member.role,
        created_at=db_member.created_at,
        updated_at=db_member.updated_at
    )
    return response


@router.get("/departments/{department_id}/members", response_model=List[DepartmentMemberResponse])
def get_department_members(
    department_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """获取部门成员列表"""
    accessible_ids = get_accessible_department_ids(db, current_user)
    
    if accessible_ids is not None and department_id not in accessible_ids:
        log_access_denied(current_user, "部门成员", department_id, "用户无权访问该部门成员列表")
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="您没有权限访问该部门"
        )
    
    # 验证部门是否存在
    department = db.query(Department).filter(Department.id == department_id).first()
    if not department:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="部门不存在"
        )
    
    # 获取部门成员
    members = db.query(DepartmentMember).filter(DepartmentMember.department_id == department_id).all()
    
    # 构建响应
    response = []
    for member in members:
        user = db.query(User).filter(User.id == member.user_id).first()
        if user:
            response.append(DepartmentMemberResponse(
                id=member.id,
                user_id=member.user_id,
                username=user.username,
                department_id=member.department_id,
                department_name=department.name,
                role=member.role,
                created_at=member.created_at,
                updated_at=member.updated_at
            ))
    
    return response


@router.put("/department-members/{member_id}", response_model=DepartmentMemberResponse)
def update_department_member(
    member_id: int,
    member_update: DepartmentMemberUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """更新部门成员。普通管理员可管理自己部门树内的成员。"""
    # 验证成员是否存在
    db_member = db.query(DepartmentMember).filter(DepartmentMember.id == member_id).first()
    if not db_member:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="部门成员不存在"
        )
    ensure_department_member_operation_scope(db, current_user, db_member)
    
    if member_update.role is None and member_update.department_id is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="未提供需要更新的字段"
        )

    old_role = db_member.role

    if member_update.role is not None:
        if member_update.role == "leader":
            existing_leader = db.query(DepartmentMember).filter(
                DepartmentMember.department_id == db_member.department_id,
                DepartmentMember.role == "leader"
            ).first()
            if existing_leader and existing_leader.id != member_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="该部门已经有组长"
                )

    if member_update.department_id is not None:
        target_department = db.query(Department).filter(
            Department.id == member_update.department_id
        ).first()
        if not target_department:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="目标部门不存在"
            )

        allowed, error_msg = can_move_member_between_departments(
            db, current_user, db_member, member_update.department_id
        )
        if not allowed:
            log_access_denied(current_user, "部门成员归属", db_member.department_id, error_msg)
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=error_msg
            )

        db_member.department_id = member_update.department_id

    if member_update.role is not None:
        db_member.role = member_update.role
    
    db.commit()
    db.refresh(db_member)
    
    # 构建响应
    user = db.query(User).filter(User.id == db_member.user_id).first()
    department = db.query(Department).filter(Department.id == db_member.department_id).first()
    
    logger.info(f"更新部门成员角色: 用户{current_user.username}将部门{department.name}的成员{user.username}角色从{old_role}改为{db_member.role}")
    
    response = DepartmentMemberResponse(
        id=db_member.id,
        user_id=db_member.user_id,
        username=user.username if user else "",
        department_id=db_member.department_id,
        department_name=department.name if department else "",
        role=db_member.role,
        created_at=db_member.created_at,
        updated_at=db_member.updated_at
    )
    return response


@router.delete("/department-members/{member_id}", response_model=Message)
def remove_department_member(
    member_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """移除部门成员。普通管理员可管理所属部门及下级部门的成员。"""
    # 验证成员是否存在
    db_member = db.query(DepartmentMember).filter(DepartmentMember.id == member_id).first()
    if not db_member:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="部门成员不存在"
        )
    ensure_department_member_operation_scope(db, current_user, db_member)
    
    # 不允许移除自己
    if db_member.user_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="不能移除自己作为部门成员"
        )
    
    # 获取信息用于日志
    user = db.query(User).filter(User.id == db_member.user_id).first()
    department = db.query(Department).filter(Department.id == db_member.department_id).first()
    
    # 删除成员
    db.delete(db_member)
    db.commit()
    
    logger.info(f"移除部门成员: 用户{current_user.username}将用户{user.username if user else '未知'}从部门{department.name if department else '未知'}移除")
    
    return Message(message="成员移除成功")


# ============ 项目管理 ============

@router.post("/projects", response_model=ProjectResponse, status_code=status.HTTP_201_CREATED)
def create_project(
    project: ProjectCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """创建项目"""
    department_ids = list(dict.fromkeys(project.department_ids or []))

    allowed, error_msg = validate_project_assignment_scope(
        db,
        current_user,
        department_ids,
        project.is_public,
    )
    if not allowed:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=error_msg
        )

    for department_id in department_ids:
        department = db.query(Department).filter(Department.id == department_id).first()
        if not department:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"部门不存在: {department_id}"
            )

    # 创建项目
    db_project = Project(
        name=project.name,
        description=project.description,
        is_public=project.is_public
    )
    db.add(db_project)
    db.commit()
    db.refresh(db_project)

    # 绑定部门（如果是私有项目）
    if not project.is_public and department_ids:
        for department_id in department_ids:
            # 创建项目-部门关联
            project_department = ProjectDepartment(
                project_id=db_project.id,
                department_id=department_id
            )
            db.add(project_department)

    db.commit()
    return build_project_detail_response(db, db_project)


@router.get("/projects/by-name/{project_name}", response_model=ProjectDetailResponse)
def get_project_by_name(
    project_name: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """通过项目名称获取项目详情"""
    # 验证项目是否存在
    project = db.query(Project).filter(Project.name == project_name).first()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="项目不存在"
        )

    allowed, error_msg = can_manage_org_project(db, current_user, project)
    if not allowed:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=error_msg
        )

    return build_project_detail_response(db, project)


@router.get("/projects", response_model=List[ProjectDetailResponse])
def get_projects(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """获取项目列表"""
    projects = db.query(Project).all()
    
    result = []
    for project in projects:
        allowed, _ = can_manage_org_project(db, current_user, project)
        if allowed:
            result.append(build_project_detail_response(db, project))
    
    return result


@router.get("/projects/{project_id}", response_model=ProjectDetailResponse)
def get_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """获取项目详情"""
    # 验证项目是否存在
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="项目不存在"
        )
    
    allowed, error_msg = can_manage_org_project(db, current_user, project)
    if not allowed:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=error_msg
        )

    return build_project_detail_response(db, project)


@router.put("/projects/{project_id}", response_model=ProjectResponse)
def update_project(
    project_id: int,
    project_update: ProjectUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """更新项目信息"""
    # 验证项目是否存在
    db_project = db.query(Project).filter(Project.id == project_id).first()
    if not db_project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="项目不存在"
        )

    allowed, error_msg = can_manage_org_project(db, current_user, db_project)
    if not allowed:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=error_msg
        )

    if is_ordinary_admin(current_user) and project_update.is_public:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="普通管理员不能将项目设置为公开项目"
        )
    
    # 更新项目信息
    if project_update.name is not None:
        db_project.name = project_update.name
    if project_update.description is not None:
        db_project.description = project_update.description
    if project_update.is_public is not None:
        db_project.is_public = project_update.is_public
    
    db.commit()
    db.refresh(db_project)
    return db_project


@router.delete("/projects/{project_id}", response_model=Message)
def delete_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """删除项目"""
    # 验证项目是否存在
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="项目不存在"
        )

    allowed, error_msg = can_manage_org_project(db, current_user, project)
    if not allowed:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=error_msg
        )
    
    # 删除项目-部门关联
    db.query(ProjectDepartment).filter(ProjectDepartment.project_id == project_id).delete()
    
    # 删除项目
    db.delete(project)
    db.commit()
    
    return Message(message="项目删除成功")
