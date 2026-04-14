"""用户路由"""

import base64
import csv
import io
import secrets
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException, Response, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session, selectinload

from app.auth import cleanup_expired_sessions, get_password_hash, verify_password
from app.database import get_db
from app.dependencies import get_current_super_admin, get_current_user, get_current_user_management_user
from app.model import Department, DepartmentMember, Role, User, UserSession
from app.rbac import (
    ensure_platform_roles_seeded,
    filter_non_platform_roles,
    get_primary_platform_role,
    normalize_role,
    normalize_role_name,
    set_user_platform_role,
)
from app.schema import (
    ChangeOwnPasswordRequest,
    ChangePasswordRequest,
    Message,
    OnlineUserResponse,
    PlatformRoleResponse,
    PlatformRoleUpdateRequest,
    UserCreate,
    UserDetailResponse,
    UserImportCommitResponse,
    UserImportNormalizedRow,
    UserImportPreviewResponse,
    UserImportRequest,
    UserImportRowResult,
    UserResponse,
    UserRoleBindRequest,
    UserRoleResponse,
    UserUpdate,
)

router = APIRouter(tags=["用户管理"], prefix="/users")

USER_IMPORT_TEMPLATE_FILENAME = "secflow-user-import-template.xlsx"
USER_IMPORT_REQUIRED_HEADERS = {"username"}
USER_IMPORT_ALLOWED_HEADERS = {
    "username",
    "password",
    "platform_role",
    "role_names",
    "department_name",
    "department_role",
    "is_active",
}
USER_IMPORT_ALLOWED_DEPARTMENT_ROLES = {"leader", "vice_leader", "member"}


def _normalize_csv_text(content: str) -> str:
    return (content or "").replace("\r\n", "\n").replace("\r", "\n").lstrip("\ufeff")


def _split_role_names(raw_value: str) -> List[str]:
    if not raw_value:
        return []

    normalized = raw_value
    for separator in ["，", ";", "；", "|", "/", "\\"]:
        normalized = normalized.replace(separator, ",")
    return [item.strip() for item in normalized.split(",") if item.strip()]


def _parse_bool(raw_value: str, default: bool = True) -> bool:
    if raw_value is None:
        return default

    value = str(raw_value).strip().lower()
    if not value:
        return default
    if value in {"true", "1", "yes", "y", "是", "启用", "active"}:
        return True
    if value in {"false", "0", "no", "n", "否", "禁用", "disabled"}:
        return False
    raise ValueError("is_active 仅支持 true/false、1/0、yes/no")


def _load_import_rows(csv_content: str) -> Tuple[List[Tuple[int, Dict[str, str]]], List[str]]:
    normalized = _normalize_csv_text(csv_content)
    if not normalized.strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV 内容不能为空"
        )

    reader = csv.DictReader(io.StringIO(normalized))
    headers = [header.strip() for header in (reader.fieldnames or []) if header and header.strip()]
    if not headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV 缺少表头"
        )

    missing_headers = USER_IMPORT_REQUIRED_HEADERS.difference(headers)
    if missing_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"CSV 缺少必填列: {', '.join(sorted(missing_headers))}"
        )

    unsupported_headers = [header for header in headers if header not in USER_IMPORT_ALLOWED_HEADERS]
    if unsupported_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"CSV 存在不支持的列: {', '.join(unsupported_headers)}"
        )

    rows: List[Tuple[int, Dict[str, str]]] = []
    for row_no, row in enumerate(reader, start=2):
        normalized_row = {
            key.strip(): (value.strip() if isinstance(value, str) else "")
            for key, value in (row or {}).items()
            if key is not None and key.strip()
        }
        if not any(normalized_row.values()):
            continue
        rows.append((row_no, normalized_row))

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV 不包含有效数据行"
        )

    return rows, headers


def _load_import_rows_from_excel(file_bytes: bytes) -> Tuple[List[Tuple[int, Dict[str, str]]], List[str]]:
    try:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Excel 文件解析失败: {exc}"
        ) from exc

    sheet = workbook.active
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [str(cell).strip() for cell in (header_cells or []) if str(cell or "").strip()]
    if not headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel 缺少表头"
        )

    missing_headers = USER_IMPORT_REQUIRED_HEADERS.difference(headers)
    if missing_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Excel 缺少必填列: {', '.join(sorted(missing_headers))}"
        )

    unsupported_headers = [header for header in headers if header not in USER_IMPORT_ALLOWED_HEADERS]
    if unsupported_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Excel 存在不支持的列: {', '.join(unsupported_headers)}"
        )

    rows: List[Tuple[int, Dict[str, str]]] = []
    for row_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        normalized_row: Dict[str, str] = {}
        has_value = False
        for index, header in enumerate(headers):
            cell_value = values[index] if values and index < len(values) else ""
            value = str(cell_value).strip() if cell_value is not None else ""
            normalized_row[header] = value
            if value:
                has_value = True
        if not has_value:
            continue
        rows.append((row_no, normalized_row))

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel 不包含有效数据行"
        )

    return rows, headers


def _load_import_payload_rows(request: UserImportRequest) -> Tuple[List[Tuple[int, Dict[str, str]]], List[str]]:
    if request.file_content_base64:
        if not request.filename:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="导入文件缺少文件名"
            )
        try:
            file_bytes = base64.b64decode(request.file_content_base64)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"导入文件内容解码失败: {exc}"
            ) from exc

        suffix = Path(request.filename).suffix.lower()
        if suffix == ".csv":
            try:
                csv_content = file_bytes.decode("utf-8-sig")
            except UnicodeDecodeError:
                csv_content = file_bytes.decode("gbk")
            return _load_import_rows(csv_content)
        if suffix == ".xlsx":
            return _load_import_rows_from_excel(file_bytes)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="当前仅支持导入 .xlsx 或 .csv 文件"
        )

    if request.csv_content is not None:
        return _load_import_rows(request.csv_content)

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="请上传 Excel/CSV 文件，或提供 CSV 内容"
    )


def _build_user_import_template_workbook() -> bytes:
    workbook = Workbook()
    template_sheet = workbook.active
    template_sheet.title = "用户导入模板"
    template_sheet.append([
        "username",
        "password",
        "platform_role",
        "role_names",
        "department_name",
        "department_role",
        "is_active",
    ])
    template_sheet.append([
        "zhangsan",
        "",
        "ordinary_user",
        "审计员",
        "安全运营部",
        "member",
        "true",
    ])
    template_sheet.append([
        "lisi",
        "TempPass123!",
        "ordinary_admin",
        "",
        "攻防实验室",
        "leader",
        "true",
    ])
    template_sheet.freeze_panes = "A2"
    for column, width in {
        "A": 20,
        "B": 20,
        "C": 18,
        "D": 28,
        "E": 22,
        "F": 18,
        "G": 14,
    }.items():
        template_sheet.column_dimensions[column].width = width
    for cell in template_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="2563EB")

    guide_sheet = workbook.create_sheet("填写说明")
    guide_sheet.column_dimensions["A"].width = 24
    guide_sheet.column_dimensions["B"].width = 96
    guide_sheet.append(["字段", "说明"])
    guide_sheet.append(["username", "必填，填写要创建的用户名，系统内不能重复"])
    guide_sheet.append(["password", "可选，留空则系统自动生成初始密码，并在导入结果里仅展示一次"])
    guide_sheet.append(["platform_role", "可选，填写 ordinary_user 或 ordinary_admin；留空默认 ordinary_user"])
    guide_sheet.append(["role_names", "可选，填写系统中已存在的普通角色名；多个角色可用逗号分隔"])
    guide_sheet.append(["department_name", "可选，填写系统中已存在的部门名称"])
    guide_sheet.append(["department_role", "可选，填写 member / vice_leader / leader；若填了部门但未填角色，默认 member"])
    guide_sheet.append(["is_active", "可选，填写 true/false、1/0、yes/no；留空默认 true"])
    guide_sheet.append(["填写建议", "通常只要填写 username，其余按需补充；先下载模板，再直接覆盖示例数据即可"])
    for cell in guide_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="0F766E")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _validate_import_row(
    db: Session,
    row_no: int,
    raw_row: Dict[str, str],
    file_username_counts: Dict[str, int],
    pending_leader_counts: Dict[int, int],
    department_cache: Dict[str, Department],
    role_cache: Dict[str, Role],
    existing_usernames: Optional[set[str]] = None,
    existing_leader_department_ids: Optional[set[int]] = None,
) -> UserImportRowResult:
    messages: List[str] = []
    username = (raw_row.get("username") or "").strip()
    password = (raw_row.get("password") or "").strip()
    platform_role_raw = (raw_row.get("platform_role") or "").strip()
    role_names = _split_role_names(raw_row.get("role_names") or "")
    department_name = (raw_row.get("department_name") or "").strip() or None
    department_role = (raw_row.get("department_role") or "").strip().lower() or None

    if not username:
        messages.append("用户名不能为空")
    elif len(username) > 100:
        messages.append("用户名长度不能超过100个字符")
    elif file_username_counts.get(username, 0) > 1:
        messages.append("CSV 文件内用户名重复")
    elif existing_usernames is not None and username in existing_usernames:
        messages.append("用户名已存在")
    elif existing_usernames is None and db.query(User).filter(User.username == username).first():
        messages.append("用户名已存在")

    if password and len(password) < 6:
        messages.append("密码长度至少6位")

    platform_role = normalize_role_name(platform_role_raw) or "ordinary_user"
    if platform_role not in {"ordinary_admin", "ordinary_user"}:
        messages.append("platform_role 仅支持 ordinary_admin 或 ordinary_user")

    resolved_roles: List[Role] = []
    for role_name in role_names:
        cached_role = role_cache.get(role_name)
        if cached_role is None:
            cached_role = db.query(Role).filter(Role.name == role_name).first()
            if cached_role:
                role_cache[role_name] = cached_role
        if cached_role is None:
            messages.append(f"角色不存在: {role_name}")
            continue
        if normalize_role(cached_role) in {"super_admin", "ordinary_admin", "ordinary_user"}:
            messages.append(f"role_names 不允许包含平台保留角色: {role_name}")
            continue
        resolved_roles.append(cached_role)

    resolved_department = None
    if department_name:
        resolved_department = department_cache.get(department_name)
        if resolved_department is None:
            resolved_department = db.query(Department).filter(Department.name == department_name).first()
            if resolved_department:
                department_cache[department_name] = resolved_department
        if resolved_department is None:
            messages.append(f"部门不存在: {department_name}")

    if department_role and department_role not in USER_IMPORT_ALLOWED_DEPARTMENT_ROLES:
        messages.append("department_role 仅支持 leader、vice_leader、member")

    if department_name and not department_role:
        department_role = "member"
    if department_role and not department_name:
        messages.append("填写 department_role 时必须同时提供 department_name")

    if resolved_department and department_role == "leader":
        has_existing_leader = (
            resolved_department.id in existing_leader_department_ids
            if existing_leader_department_ids is not None
            else db.query(DepartmentMember).filter(
                DepartmentMember.department_id == resolved_department.id,
                DepartmentMember.role == "leader"
            ).first() is not None
        )
        if has_existing_leader:
            messages.append(f"部门已有组长: {department_name}")
        if pending_leader_counts.get(resolved_department.id, 0) > 1:
            messages.append(f"CSV 中部门组长重复: {department_name}")

    try:
        is_active = _parse_bool(raw_row.get("is_active") or "", default=True)
    except ValueError as exc:
        messages.append(str(exc))
        is_active = True

    status_value = "valid" if not messages else "error"
    return UserImportRowResult(
        row_no=row_no,
        username=username,
        status=status_value,
        messages=messages,
        normalized=UserImportNormalizedRow(
            username=username,
            password_provided=bool(password),
            platform_role=platform_role,
            role_names=[role.name for role in resolved_roles],
            department_name=department_name,
            department_role=department_role,
            is_active=is_active,
        )
    )


def _preview_import_rows(db: Session, request: UserImportRequest) -> UserImportPreviewResponse:
    rows, _headers = _load_import_payload_rows(request)
    ensure_platform_roles_seeded(db)

    file_username_counts: Dict[str, int] = {}
    pending_leader_counts: Dict[int, int] = {}
    department_cache: Dict[str, Department] = {}
    role_cache: Dict[str, Role] = {}

    usernames = sorted({
        (raw_row.get("username") or "").strip()
        for _row_no, raw_row in rows
        if (raw_row.get("username") or "").strip()
    })
    existing_usernames = {
        username
        for username, in db.query(User.username).filter(User.username.in_(usernames)).all()
    }

    department_names = sorted({
        (raw_row.get("department_name") or "").strip()
        for _row_no, raw_row in rows
        if (raw_row.get("department_name") or "").strip()
    })
    if department_names:
        department_cache.update({
            department.name: department
            for department in db.query(Department).filter(Department.name.in_(department_names)).all()
        })

    role_names = sorted({
        role_name
        for _row_no, raw_row in rows
        for role_name in _split_role_names(raw_row.get("role_names") or "")
    })
    if role_names:
        role_cache.update({
            role.name: role
            for role in db.query(Role).filter(Role.name.in_(role_names)).all()
        })

    leader_department_ids = [
        department_cache[department_name].id
        for _row_no, raw_row in rows
        for department_name in [(raw_row.get("department_name") or "").strip()]
        if department_name
        and (raw_row.get("department_role") or "").strip().lower() == "leader"
        and department_name in department_cache
    ]
    existing_leader_department_ids = {
        department_id
        for department_id, in db.query(DepartmentMember.department_id).filter(
            DepartmentMember.department_id.in_(leader_department_ids),
            DepartmentMember.role == "leader"
        ).all()
    }

    for _row_no, raw_row in rows:
        username = (raw_row.get("username") or "").strip()
        if username:
            file_username_counts[username] = file_username_counts.get(username, 0) + 1

        department_name = (raw_row.get("department_name") or "").strip()
        department_role = (raw_row.get("department_role") or "").strip().lower()
        if department_name and department_role == "leader":
            department = department_cache.get(department_name)
            if department:
                pending_leader_counts[department.id] = pending_leader_counts.get(department.id, 0) + 1

    results = [
        _validate_import_row(
            db=db,
            row_no=row_no,
            raw_row=raw_row,
            file_username_counts=file_username_counts,
            pending_leader_counts=pending_leader_counts,
            department_cache=department_cache,
            role_cache=role_cache,
            existing_usernames=existing_usernames,
            existing_leader_department_ids=existing_leader_department_ids,
        )
        for row_no, raw_row in rows
    ]

    valid_rows = sum(1 for item in results if item.status == "valid")
    return UserImportPreviewResponse(
        total_rows=len(results),
        valid_rows=valid_rows,
        error_rows=len(results) - valid_rows,
        rows=results,
    )


def _generate_initial_password() -> str:
    return secrets.token_urlsafe(10)


def _get_primary_department_membership(db: Session, user_id: int) -> Optional[DepartmentMember]:
    return db.query(DepartmentMember).filter(
        DepartmentMember.user_id == user_id
    ).order_by(DepartmentMember.id.asc()).first()


def _build_primary_membership_map(db: Session, user_ids: List[int]) -> Dict[int, DepartmentMember]:
    unique_user_ids = [user_id for user_id in set(user_ids) if user_id is not None]
    if not unique_user_ids:
        return {}

    memberships = db.query(DepartmentMember).filter(
        DepartmentMember.user_id.in_(unique_user_ids)
    ).order_by(
        DepartmentMember.user_id.asc(),
        DepartmentMember.id.asc()
    ).all()

    membership_map: Dict[int, DepartmentMember] = {}
    for membership in memberships:
        if membership.user_id not in membership_map:
            membership_map[membership.user_id] = membership
    return membership_map


def _build_department_map(db: Session, department_ids: List[int]) -> Dict[int, Department]:
    unique_department_ids = [department_id for department_id in set(department_ids) if department_id is not None]
    if not unique_department_ids:
        return {}
    departments = db.query(Department).filter(Department.id.in_(unique_department_ids)).all()
    return {department.id: department for department in departments}


def _build_user_response(
    db: Session,
    user: User,
    membership_map: Optional[Dict[int, DepartmentMember]] = None,
    department_map: Optional[Dict[int, Department]] = None,
) -> UserResponse:
    membership = membership_map.get(user.id) if membership_map is not None else _get_primary_department_membership(db, user.id)
    department = department_map.get(membership.department_id) if (membership and department_map is not None) else None
    if membership and department is None and department_map is None:
        department = db.query(Department).filter(Department.id == membership.department_id).first()

    return UserResponse(
        id=user.id,
        username=user.username,
        is_active=user.is_active,
        created_at=user.created_at,
        updated_at=user.updated_at,
        role=user.get_all_role_names(),
        platform_role=get_primary_platform_role(user),
        department_member_id=membership.id if membership else None,
        department_id=membership.department_id if membership else None,
        department_name=department.name if department else None,
    )


def _build_user_responses(db: Session, users: List[User]) -> List[UserResponse]:
    membership_map = _build_primary_membership_map(db, [user.id for user in users])
    department_map = _build_department_map(
        db,
        [membership.department_id for membership in membership_map.values()]
    )
    return [
        _build_user_response(
            db,
            user,
            membership_map=membership_map,
            department_map=department_map,
        )
        for user in users
    ]


def _validate_role_ids_exist(db: Session, role_ids: List[int]) -> List[Role]:
    roles = db.query(Role).filter(Role.id.in_(role_ids)).all()
    if len(roles) != len(role_ids):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="部分角色不存在"
        )
    return roles


@router.get("/user_list", response_model=List[UserResponse])
def list_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_management_user)
):
    """获取用户列表。用户管理范围内的管理员可访问。"""
    ensure_platform_roles_seeded(db)
    users = db.query(User).options(
        selectinload(User.roles)
    ).order_by(User.id.asc()).all()
    return _build_user_responses(db, users)


@router.get("/{user_id}", response_model=UserDetailResponse)
def get_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """获取单个用户。仅超级管理员可访问。"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    response = _build_user_response(db, user)
    return UserDetailResponse(
        **response.model_dump(),
        role_ids=[role.id for role in user.roles]
    )


@router.post("", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def create_user(
    user_data: UserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """创建用户。未绑定固定管理角色时即为普通用户。"""
    existing = db.query(User).filter(User.username == user_data.username).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="用户名已存在"
        )

    ensure_platform_roles_seeded(db)

    user = User(
        username=user_data.username,
        hashed_password=get_password_hash(user_data.password)
    )
    db.add(user)
    db.flush()

    extra_roles: List[Role] = []
    if user_data.role_ids:
        roles = _validate_role_ids_exist(db, user_data.role_ids)
        extra_roles = filter_non_platform_roles(roles)

    user.roles = list(extra_roles)

    db.commit()
    db.refresh(user)
    return _build_user_response(db, user)


@router.get("/import/template")
def download_user_import_template(
    current_user: User = Depends(get_current_super_admin)
):
    """下载用户导入 Excel 模板。"""
    workbook_bytes = _build_user_import_template_workbook()
    return Response(
        content=workbook_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{USER_IMPORT_TEMPLATE_FILENAME}"'
        }
    )


@router.post("/import/preview", response_model=UserImportPreviewResponse)
def preview_user_import(
    request: UserImportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """预校验用户导入文件。"""
    return _preview_import_rows(db, request)


@router.post("/import/commit", response_model=UserImportCommitResponse)
def commit_user_import(
    request: UserImportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """执行用户导入。按行提交，单行失败不影响整批。"""
    preview = _preview_import_rows(db, request)
    raw_rows, _headers = _load_import_payload_rows(request)
    raw_row_map = {row_no: item for row_no, item in raw_rows}
    results: List[UserImportRowResult] = []

    for row in preview.rows:
        if row.status != "valid" or row.normalized is None:
            results.append(row)
            continue

        normalized = row.normalized
        raw_row = raw_row_map.get(row.row_no, {})
        password = (raw_row.get("password") or "").strip() or _generate_initial_password()

        try:
            user = User(
                username=normalized.username,
                hashed_password=get_password_hash(password),
                is_active=normalized.is_active,
            )
            db.add(user)
            db.flush()

            if normalized.platform_role == "ordinary_admin":
                set_user_platform_role(db, user, normalized.platform_role)

            if normalized.role_names:
                roles = db.query(Role).filter(Role.name.in_(normalized.role_names)).all()
                user.roles = list(user.roles) + filter_non_platform_roles(roles)

            if normalized.department_name:
                department = db.query(Department).filter(Department.name == normalized.department_name).first()
                if not department:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"部门不存在: {normalized.department_name}"
                    )

                if normalized.department_role == "leader":
                    existing_leader = db.query(DepartmentMember).filter(
                        DepartmentMember.department_id == department.id,
                        DepartmentMember.role == "leader"
                    ).first()
                    if existing_leader:
                        raise HTTPException(
                            status_code=status.HTTP_400_BAD_REQUEST,
                            detail=f"部门已有组长: {normalized.department_name}"
                        )

                db.add(DepartmentMember(
                    user_id=user.id,
                    department_id=department.id,
                    role=normalized.department_role or "member"
                ))

            user.updated_at = datetime.utcnow()
            db.commit()
            db.refresh(user)
            results.append(UserImportRowResult(
                row_no=row.row_no,
                username=normalized.username,
                status="success",
                messages=["导入成功"],
                normalized=normalized,
                generated_password=None if raw_row.get("password") else password,
                user_id=user.id,
            ))
        except HTTPException as exc:
            db.rollback()
            results.append(UserImportRowResult(
                row_no=row.row_no,
                username=normalized.username,
                status="error",
                messages=[str(exc.detail)],
                normalized=normalized,
            ))
        except Exception as exc:
            db.rollback()
            results.append(UserImportRowResult(
                row_no=row.row_no,
                username=normalized.username,
                status="error",
                messages=[f"导入失败: {str(exc)}"],
                normalized=normalized,
            ))

    success_rows = sum(1 for item in results if item.status == "success")
    return UserImportCommitResponse(
        total_rows=len(results),
        success_rows=success_rows,
        failed_rows=len(results) - success_rows,
        rows=results,
    )


@router.put("/{user_id}", response_model=UserResponse)
def update_user(
    user_id: int,
    user_data: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """更新用户基础信息。仅超级管理员可访问。"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    if user_data.username:
        existing = db.query(User).filter(
            User.username == user_data.username,
            User.id != user_id
        ).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="用户名已存在"
            )
        user.username = user_data.username

    if user_data.password:
        user.hashed_password = get_password_hash(user_data.password)

    if user_data.is_active is not None:
        user.is_active = user_data.is_active

    user.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(user)

    return _build_user_response(db, user)


@router.delete("/{user_id}", response_model=Message)
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """删除用户。仅超级管理员可访问。"""
    if current_user.id == user_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="不能删除当前登录的超级管理员"
        )

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    db.delete(user)
    db.commit()

    return {"message": "用户已删除"}


@router.get("/{user_id}/role", response_model=UserRoleResponse)
def get_user_roles(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """获取用户的原始角色列表。仅超级管理员可访问。"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    return UserRoleResponse(
        user_id=user.id,
        role_ids=[role.id for role in user.roles],
        role_names=[role.name for role in user.roles]
    )


@router.put("/{user_id}/role", response_model=UserRoleResponse)
def bind_user_role(
    user_id: int,
    role_data: UserRoleBindRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """覆盖用户角色。保留接口兼容性，仅超级管理员可访问。"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    roles = _validate_role_ids_exist(db, role_data.role_ids)
    normalized_role_names = {
        normalize_role(role)
        for role in roles
    }
    if "super_admin" in normalized_role_names:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="超级管理员角色为保留角色，不能通过该接口直接分配"
        )

    user.roles = roles
    user.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(user)

    return UserRoleResponse(
        user_id=user.id,
        role_ids=[role.id for role in user.roles],
        role_names=[role.name for role in user.roles]
    )


@router.post("/{user_id}/role/add", response_model=UserRoleResponse)
def add_user_role(
    user_id: int,
    role_data: UserRoleBindRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """增量添加角色。仅超级管理员可访问。"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    roles = _validate_role_ids_exist(db, role_data.role_ids)
    if any(normalize_role(role) == "super_admin" for role in roles):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="超级管理员角色为保留角色，不能通过该接口直接分配"
        )

    all_roles = list(user.roles)
    known_ids = {role.id for role in all_roles}
    for role in roles:
        if role.id not in known_ids:
            known_ids.add(role.id)
            all_roles.append(role)

    user.roles = all_roles
    user.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(user)

    return UserRoleResponse(
        user_id=user.id,
        role_ids=[role.id for role in user.roles],
        role_names=[role.name for role in user.roles]
    )


@router.delete("/{user_id}/role", response_model=UserRoleResponse)
def remove_user_role(
    user_id: int,
    role_data: UserRoleBindRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """移除角色。至少保留一个平台角色。"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    user.roles = [role for role in user.roles if role.id not in set(role_data.role_ids)]
    user.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(user)

    return UserRoleResponse(
        user_id=user.id,
        role_ids=[role.id for role in user.roles],
        role_names=[role.name for role in user.roles]
    )


@router.get("/{user_id}/platform-role", response_model=PlatformRoleResponse)
def get_user_platform_role(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """获取用户的平台角色。"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    return PlatformRoleResponse(
        user_id=user.id,
        platform_role=get_primary_platform_role(user),
        role_names=user.get_all_role_names()
    )


@router.put("/{user_id}/platform-role", response_model=PlatformRoleResponse)
def update_user_platform_role(
    user_id: int,
    request: PlatformRoleUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """仅允许超级管理员在普通管理员与普通用户之间切换。"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    if user.id == current_user.id or user.id == 1:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="当前超级管理员账号的角色不允许在此页面中修改"
        )

    set_user_platform_role(db, user, request.role_name)
    user.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(user)

    return PlatformRoleResponse(
        user_id=user.id,
        platform_role=get_primary_platform_role(user),
        role_names=user.get_all_role_names()
    )


@router.post("/{user_id}/password", response_model=Message)
def change_password(
    user_id: int,
    request: ChangePasswordRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """
    超级管理员修改指定用户的密码。

    old_password 用于校验当前超级管理员自己的密码，避免高权限操作被误用。
    """
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    if not verify_password(request.old_password, current_user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="超级管理员密码验证失败"
        )

    if not request.new_password or len(request.new_password) < 6:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="新密码不能为空且长度至少6位"
        )

    user.hashed_password = get_password_hash(request.new_password)
    user.updated_at = datetime.utcnow()

    db.commit()

    return {"message": "密码已修改"}


@router.post("/password/self", response_model=Message)
def change_own_password(
    request: ChangeOwnPasswordRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    当前登录用户修改自己的密码。
    """
    if not verify_password(request.old_password, current_user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="旧密码错误"
        )

    if not request.new_password or len(request.new_password) < 6:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="新密码不能为空且长度至少6位"
        )

    current_user.hashed_password = get_password_hash(request.new_password)
    current_user.updated_at = datetime.utcnow()

    db.commit()

    return {"message": "密码已修改"}


@router.get("/sessions/online", response_model=List[OnlineUserResponse])
def get_online_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """
    获取当前在线用户列表。仅超级管理员可访问。
    """
    cleanup_expired_sessions(db)

    now = datetime.utcnow()
    active_sessions = db.query(UserSession).filter(
        UserSession.status == "active",
        UserSession.expires_at > now
    ).order_by(UserSession.last_active_at.desc()).all()

    online_users = []
    seen_user_ids = set()

    for session in active_sessions:
        if session.user_id in seen_user_ids:
            continue
        seen_user_ids.add(session.user_id)

        user = session.user
        online_users.append(OnlineUserResponse(
            user_id=user.id,
            username=user.username,
            role=user.get_all_role_names(),
            ip_address=session.ip_address,
            user_agent=session.user_agent,
            login_at=session.created_at,
            last_active_at=session.last_active_at
        ))

    return online_users


@router.get("/{user_id}/sessions", response_model=List[dict])
def get_user_sessions(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """
    获取指定用户的所有会话。仅超级管理员可访问。
    """
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    cleanup_expired_sessions(db)

    sessions = db.query(UserSession).filter(
        UserSession.user_id == user_id
    ).order_by(UserSession.created_at.desc()).all()

    return [
        {
            "id": s.id,
            "token_jti": s.token_jti,
            "ip_address": s.ip_address,
            "user_agent": s.user_agent,
            "status": s.status,
            "created_at": s.created_at.isoformat() if s.created_at else None,
            "last_active_at": s.last_active_at.isoformat() if s.last_active_at else None,
            "expires_at": s.expires_at.isoformat() if s.expires_at else None
        }
        for s in sessions
    ]


@router.delete("/{user_id}/sessions", response_model=Message)
def revoke_user_sessions(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_super_admin)
):
    """
    撤销指定用户的所有会话（踢下线）。仅超级管理员可访问。
    """
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    active_sessions = db.query(UserSession).filter(
        UserSession.user_id == user_id,
        UserSession.status == "active"
    ).all()

    for session in active_sessions:
        session.status = "revoked"

    db.commit()

    return {"message": f"已撤销用户 {user.username} 的所有会话"}
